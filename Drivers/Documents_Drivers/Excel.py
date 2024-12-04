from openpyxl import load_workbook
import pandas as pd
import json  
import tempfile
import os
import shutil
import csv
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

class Excel:

    def __init__(self, caminho_arquivo=False, conteudo_binario = False, read_only=False):
        if caminho_arquivo:
            self.caminho_arquivo = caminho_arquivo
            self.wb = load_workbook(caminho_arquivo,read_only = read_only)
        if conteudo_binario:
            self.conteudo_binario = conteudo_binario
        self.safe_path = False

    def Gerar_Dicionario_de_Grupos_de_Colunas(self, nome_aba, CaminhoArquivoJson=False, NomeGrupoPrimeiraLinha=False, 
                                                index =  False):
        df = pd.read_excel(self.caminho_arquivo, sheet_name=nome_aba)
        dados_grupos = {}
        grupo_numero = 1  
        colunas_grupo = {}
        nome_grupo = f'Grupo {grupo_numero}'  
        Ncol = 0
        for coluna in df.columns:
            if df[coluna].isnull().all():
                if colunas_grupo: 
                    grupo_df = pd.DataFrame(colunas_grupo).dropna(how="all").fillna("")
                    if index:
                        coluna_index = list(range(len(grupo_df)))
                        grupo_df_indexado = {'index': coluna_index}
                        grupo_df = pd.DataFrame(grupo_df_indexado).join(grupo_df.reset_index(drop=True))
                    dados_grupos[nome_grupo] = grupo_df.to_dict(orient="records")
                    grupo_numero += 1
                    nome_grupo = f'Grupo {grupo_numero}'
                    colunas_grupo = {} 
                    Ncol = 0
            else:
                if NomeGrupoPrimeiraLinha:
                    if Ncol == 0: nome_grupo = coluna
                    Ncol += 1
                    header = df[coluna].iloc[0]
                    dados = df[coluna].iloc[1:].reset_index(drop=True)  
                    colunas_grupo[header] = dados
                else:
                    colunas_grupo[coluna] = df[coluna]  
        if colunas_grupo:
            grupo_df = pd.DataFrame(colunas_grupo).dropna(how="all")
            if index:
                coluna_index = list(range(len(grupo_df)))
                grupo_df_indexado = {'index': coluna_index}
                grupo_df = pd.DataFrame(grupo_df_indexado).join(grupo_df.reset_index(drop=True))
            dados_grupos[nome_grupo] = grupo_df.to_dict(orient="records")
        if CaminhoArquivoJson:
            with open(CaminhoArquivoJson, "w", encoding="utf-8") as json_file:
                json.dump(dados_grupos, json_file, ensure_ascii=False, indent=4)
        return dados_grupos
    
    def abrir_editar_planilha_conteudo_binario(self, nome_arquivo):
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
            temp_file.write(self.conteudo_binario)
        temp_file_path = temp_file.name
        self.safe_path = os.path.join(os.path.dirname(temp_file_path), nome_arquivo)
        shutil.copy(temp_file_path, self.safe_path)
        os.startfile(self.safe_path)
        
    def obter_conteudo_binario_planilha_editada(self):
        conteudo_atualizado= False
        if self.safe_path and os.path.exists(self.safe_path):
            with open(self.safe_path , 'rb') as updated_file:
                conteudo_atualizado = updated_file.read()
            self.excluir_arquivo_temp_binario()
        return conteudo_atualizado

    def excluir_arquivo_temp_binario(self):
        try:
            if self.safe_path and os.path.exists(self.safe_path):
                os.remove(self.safe_path )
        except:
            print('Não foi possível excluir o arquivo')
    
    def obter_data_frame_convertendo_em_csv(self):
        sheet = self.wb.active
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".csv", mode="w", newline="", encoding="utf-8")
        with temp_file as f:
            writer = csv.writer(f)
            for row in sheet.iter_rows(values_only=True):
                writer.writerow(row)
            temp_file_path = temp_file.name
        self.wb.close()
        df = pd.read_csv(temp_file_path)
        os.remove(temp_file_path)
        return df

    def remove_bordas(self, alvo, sheet = 0):
        sheet = self.wb[sheet] if isinstance(sheet, str) else self.wb.worksheets[sheet]
        if isinstance(alvo, int):  
            for cell in sheet[alvo]:  
                cell.border = Border() 
        elif isinstance(alvo, str): 
            for row in sheet[alvo]: 
                for cell in row: 
                    cell.border = Border()  
        else:
            raise ValueError("O argumento 'alvo' deve ser um número (linha) ou uma string (intervalo)!")
    
    def mesclar_titulos(self, row_titulos, sheet = 0):
        sheet = self.wb[sheet] if isinstance(sheet, str) else self.wb.worksheets[sheet]
        group = 0 
        for cell in sheet[row_titulos + 1]:
            if cell.value is None:
                sheet.merge_cells(start_row = row_titulos, start_column = group + 1, 
                                  end_row = row_titulos, end_column = cell.column - 1)
                group = cell.column

    def auto_ajustar_largura_colunas(self, sheet=0):
        sheet = self.wb[sheet] if isinstance(sheet, str) else self.wb.worksheets[sheet]
        for column in sheet.columns:
            fez = False
            max_length = 0
            column_letter = get_column_letter(column[0].column) 
            for cell in column:
                if cell.value and cell.value.strip() != "":
                    fez = True
                    max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max_length + 5
                    sheet.column_dimensions[column_letter].width = adjusted_width
            print(fez)

    def salvar_arquivo(self, caminho_arquivo = None):
        caminho_arquivo = caminho_arquivo or self.caminho_arquivo
        if not caminho_arquivo:
            raise ValueError("O caminho do arquivo não foi especificado!")
        self.wb.save(caminho_arquivo)


    def formatacao_condicional(self, cells, operator, formula, color, color_end_gradiente = False,
                                type = "solid", sheet=0):
        sheet = self.wb[sheet] if isinstance(sheet, str) else self.wb.worksheets[sheet]
        color_end = color_end_gradiente if color_end_gradiente else color
        green_fill = PatternFill(start_color=color, end_color=color_end, fill_type="solid")
        rule = CellIsRule(operator=operator, formula=formula, fill=green_fill)
        sheet.conditional_formatting.add(cells, rule)

    def retirar_linhas_grade(self, sheet=0):
        sheet = self.wb[sheet] if isinstance(sheet, str) else self.wb.worksheets[sheet]
        sheet.sheet_view.showGridLines = False

    def definir_zoom(self, zoom, sheet=0):
        sheet = self.wb[sheet] if isinstance(sheet, str) else self.wb.worksheets[sheet]
        sheet.sheet_view.zoomScale = zoom

    def alterar_nome_aba(self, novo_nome, sheet=0):
        sheet = self.wb[sheet] if isinstance(sheet, str) else self.wb.worksheets[sheet]
        sheet.title =novo_nome